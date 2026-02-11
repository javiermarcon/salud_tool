"""Generación de Excel formateado para entrega médica."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, Side

_DIA_SEMANA: tuple[str, ...] = ("lun", "mar", "mie", "jue", "vie", "sab", "dom")

_HEADER_MAP: dict[str, str] = {
    "weekday": "Día",
    "datetime": "Fecha / Hora",
    "glucose_mg_dl": "Glucosa (mg/dL)",
    "tag": "Tag",
    "steps": "Pasos",
    "distance_m": "Distancia (m)",
    "calories_kcal": "Calorías\n(kcal)",
    "active_minutes": "Minutos\nactivos",
}


@dataclass(frozen=True)
class ExcelLayout:
    """Layout/formatting configuration for the doctor sheet."""

    sheet_name: str = "Consolidado diario"


def _weekday_label(i: object) -> str:
    """Convierte índice 0-6 (lunes-domingo) a etiqueta de 3 letras."""
    try:
        if i is None or (isinstance(i, float) and pd.isna(i)):
            return ""
        if isinstance(i, int | float):
            idx = int(i)
            return _DIA_SEMANA[idx] if 0 <= idx < 7 else ""
        return ""
    except (ValueError, TypeError):
        return ""


def _add_weekday_column(export_df: pd.DataFrame) -> pd.DataFrame:
    """Añade columna weekday (Día) a partir de date o datetime."""
    if "date" in export_df.columns:
        weekday_series = pd.to_datetime(export_df["date"]).dt.weekday
    elif "datetime" in export_df.columns:
        weekday_series = pd.to_datetime(export_df["datetime"]).dt.weekday
    else:
        return export_df
    if weekday_series.empty:
        return export_df
    export_df = export_df.copy()
    export_df["weekday"] = weekday_series.map(_weekday_label)
    cols = ["weekday"] + [c for c in export_df.columns if c != "weekday"]
    return export_df[cols]


def _prepare_datetime_and_drop_date(export_df: pd.DataFrame) -> pd.DataFrame:
    """Quita timezone de datetime y elimina columna date."""
    export_df = export_df.copy()
    if "datetime" in export_df.columns:
        export_df["datetime"] = pd.to_datetime(
            export_df["datetime"], errors="coerce"
        ).dt.tz_localize(None)
    if "date" in export_df.columns:
        export_df = export_df.drop(columns=["date"])
    return export_df


def write_doctor_xlsx(df: pd.DataFrame, out_path: Path, layout: ExcelLayout) -> None:
    """Write a formatted Excel file suitable for printing.

    Args:
        df: Consolidated daily DataFrame.
        out_path: Output path for the XLSX file.
        layout: Excel layout parameters.
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)

    export_df = df.copy()
    export_df = _add_weekday_column(export_df)
    export_df = _prepare_datetime_and_drop_date(export_df)
    export_df = export_df.rename(columns=_HEADER_MAP)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name=layout.sheet_name)
        ws = writer.book[layout.sheet_name]
        _format_sheet(ws)


def _style_header_row(ws: Any) -> None:
    """Aplica fuente negrita, alineación y borde a la fila de cabecera."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center
        cell.border = border


def _style_body_rows(ws: Any) -> None:
    """Aplica alineación, borde y altura fija a las filas de datos."""
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = center
            cell.border = border
        ws.row_dimensions[row[0].row].height = 15


def _get_header_col_index(ws: Any) -> dict[str, int]:
    """Devuelve mapa nombre de cabecera -> índice de columna (1-based)."""
    headers = [str(cell.value) for cell in ws[1]]
    return {name: idx + 1 for idx, name in enumerate(headers)}


def _apply_column_widths(ws: Any, col_index: dict[str, int]) -> None:
    """Establece anchos de columna para evitar ###."""
    widths = [
        ("Día", 6),
        ("Fecha / Hora", 18),
        ("Glucosa (mg/dL)", 14),
        ("Tag", 14),
        ("Pasos", 12),
        ("Distancia (m)", 14),
        ("Calorías\n(kcal)", 10),
        ("Minutos\nactivos", 10),
    ]
    for header, width in widths:
        idx = col_index.get(header)
        if idx is not None:
            letter = ws.cell(row=1, column=idx).column_letter
            ws.column_dimensions[letter].width = width


def _apply_number_formats(ws: Any, col_index: dict[str, int]) -> None:
    """Aplica formatos numéricos por cabecera."""
    fmt_map: dict[str, str] = {
        "Fecha / Hora": "dd/mm/yyyy hh:mm",
        "Glucosa (mg/dL)": "0.00",
        "Pasos": "#,##0",
        "Distancia (m)": "#,##0",
        "Calorías\n(kcal)": "#,##0",
        "Minutos\nactivos": "0",
    }
    for row in ws.iter_rows(min_row=2):
        for header, fmt in fmt_map.items():
            idx = col_index.get(header)
            if idx is not None:
                row[idx - 1].number_format = fmt


def _format_sheet(ws: Any) -> None:
    """Apply borders, widths and number formats to a worksheet.

    Args:
        ws: openpyxl worksheet.
    """
    _style_header_row(ws)
    _style_body_rows(ws)
    col_index = _get_header_col_index(ws)
    _apply_column_widths(ws, col_index)
    _apply_number_formats(ws, col_index)
