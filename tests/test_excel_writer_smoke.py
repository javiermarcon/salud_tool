from __future__ import annotations

from pathlib import Path
from typing import cast

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from salud_tool.excel_writer import ExcelLayout, _format_sheet, write_doctor_xlsx


def test_write_doctor_xlsx_happy_path_and_formatting(tmp_path: Path) -> None:
    """Una fila por medición: Día, Fecha / Hora, Glucosa (mg/dL), Pasos, etc."""
    df = pd.DataFrame(
        {
            "date": [
                pd.to_datetime("2025-12-15").date(),
                pd.to_datetime("2025-12-16").date(),
            ],
            "datetime": [
                pd.Timestamp("2025-12-15 08:30", tz="UTC"),
                pd.Timestamp("2025-12-16 09:45", tz="UTC"),
            ],
            "glucose_mg_dl": [105.0, 100.0],
            "steps": [1000, 2000],
            "distance_m": [500.0, 1200.0],
            "calories_kcal": [120.0, 150.0],
            "active_minutes": [22.0, 18.0],
        }
    )
    out = tmp_path / "nested" / "out.xlsx"
    write_doctor_xlsx(df, out, ExcelLayout())

    wb = load_workbook(out)
    ws = cast(Worksheet, wb[ExcelLayout().sheet_name])

    headers = [cell.value for cell in ws[1]]
    assert headers[0] == "Día"
    assert "Fecha / Hora" in headers
    assert "Glucosa (mg/dL)" in headers
    assert "Pasos" in headers
    assert "datetime" not in headers

    assert ws.cell(row=2, column=1).value == "lun"
    gluc_col = headers.index("Glucosa (mg/dL)") + 1
    assert ws.cell(row=2, column=gluc_col).value == 105.0

    assert ws.column_dimensions["A"].width == 6
    pasos_letter = get_column_letter(headers.index("Pasos") + 1)
    assert ws.column_dimensions[pasos_letter].width == 12

    pasos_cell = ws.cell(row=2, column=headers.index("Pasos") + 1)
    assert pasos_cell.number_format == "#,##0"
    gluc_cell = ws.cell(row=2, column=headers.index("Glucosa (mg/dL)") + 1)
    assert gluc_cell.number_format == "0.00"


def test_write_doctor_xlsx_with_none_date_fills_weekday_empty(tmp_path: Path) -> None:
    """Si hay fecha None, weekday se escribe vacío (no se lanza excepción)."""
    df = pd.DataFrame(
        {
            "date": [pd.to_datetime("2025-12-15").date(), None],
            "glucose_mg_dl": [100.0, 110.0],
        }
    )
    out = tmp_path / "out.xlsx"
    write_doctor_xlsx(df, out, ExcelLayout())
    wb = load_workbook(out)
    ws = wb[ExcelLayout().sheet_name]
    assert ws.cell(row=2, column=1).value == "lun"
    # Excel guarda celdas vacías como None
    assert ws.cell(row=3, column=1).value in ("", None)


def test_format_sheet_handles_missing_headers() -> None:
    wb = Workbook()
    ws = cast(Worksheet, wb.active)
    ws.append(["Solo"])
    ws.append([1])

    _format_sheet(ws)

    assert ws.cell(row=1, column=1).font.bold is True
    assert ws.cell(row=2, column=1).alignment.horizontal == "center"
